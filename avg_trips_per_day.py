"""
    The code is a Python script that reads a CSV file, calculates the number of trips for each provider type on weekdays and weekends, and outputs the result in an Excel file.
    The script imports the necessary libraries - csv, datetime, openpyxl, and get_column_letter from openpyxl.utils.
    The main function in the script is "average_trips()" that performs the following steps:
    1. Reads the CSV file "trips.csv" and stores the contents in a list called "rows".
    2. Opens an existing Excel workbook "SUMMARY_OUTPUTS.xlsx" and creates a new sheet named "question_1_part_2".
    3. Writes the header row in the new sheet.
    4. Calculates the number of trips for each provider type on weekdays and stores them in a dictionary called "weekday_counts".
    5. Calculates the number of trips for each provider type on weekends and stores them in a dictionary called "weekend_counts".
    6. Calculates the average number of trips per day for each provider type on weekdays and weekends separately, and stores them in the "weekday_row" and "weekend_row" lists.
    7. Calculates the total number of trips for each provider type and stores them in the "total_row" list.
    8. Sets the cell styles for the values in the new sheet.
    9. Saves the workbook with the new sheet.
"""

import csv
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def average_trips():
    # Read the CSV file
    with open('trips.csv') as file:
        reader = csv.DictReader(file)
        rows = list(reader)

    # Open the existing workbook and create a new sheet
    workbook = load_workbook('SUMMARY_OUTPUTS.xlsx')
    sheet = workbook.create_sheet(title='question_1_part_2')

    # Write the header row
    header = [""] + ["Primary", "Broker", "E-Hail", "Total"]
    sheet.append(header)

    # Calculate the number of trips for each provider type on weekdays
    weekday_trips = [row for row in rows if datetime.strptime(row['Tripdate'], '%Y-%m-%d').weekday() < 5]

    weekday_counts = {
        'Primary': len([trip for trip in weekday_trips if trip['ProviderType'] == 'Primary']),
        'Broker': len([trip for trip in weekday_trips if trip['ProviderType'] == 'Broker']),
        'E-Hail': len([trip for trip in weekday_trips if trip['ProviderType'] == 'E-Hail'])
    }

    weekday_row = ["Weekday (M to F)"] + [weekday_counts["Primary"]/5, weekday_counts["Broker"]/5, weekday_counts["E-Hail"]/5, sum(weekday_counts.values())/5]
    sheet.append(weekday_row)

    # Calculate the number of trips for each provider type on weekends
    weekend_trips = [row for row in rows if datetime.strptime(row['Tripdate'], '%Y-%m-%d').weekday() >= 5]

    weekend_counts = {
        'Primary': len([trip for trip in weekend_trips if trip['ProviderType'] == 'Primary']),
        'Broker': len([trip for trip in weekend_trips if trip['ProviderType'] == 'Broker']),
        'E-Hail': len([trip for trip in weekend_trips if trip['ProviderType'] == 'E-Hail'])
    }

    weekend_row = ["Weekend (Sa & Su)"] + [weekend_counts["Primary"]/2, weekend_counts["Broker"]/2, weekend_counts["E-Hail"]/2, sum(weekend_counts.values())/2]
    sheet.append(weekend_row)

    total_row = ["Total"] + [weekday_row[1]+weekend_row[1], weekday_row[2]+weekend_row[2], weekday_row[3]+weekend_row[3], weekday_row[4]+weekend_row[4]]
    sheet.append(total_row)

    # Set the cell styles
    for row in sheet.iter_rows(min_row=2, max_row=4):
        for cell in row:
            cell.number_format = '0.0'

    first_column = sheet.column_dimensions[get_column_letter(1)]
    first_column.width = 15

    # Save the workbook
    workbook.save("SUMMARY_OUTPUTS.xlsx")