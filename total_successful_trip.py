"""
    This Python script reads data from a CSV file containing information about trips, and generates summary statistics in an Excel workbook.
    The script defines two helper functions is_weekday and is_weekend to check whether a given date is a weekday or a weekend, respectively.
    The total_trips function opens the trips.csv file and iterates over its rows using a csv.DictReader object. The function reads the Tripdate,
    ProviderType, and Outcome columns from each row, and skips over any row with a ProviderType of "Test Record", "Deleted", or "Driver break record",
    or an Outcome of "Fixed Route-Exclude". For each valid row, the function increments a counter for the provider type of the trip, based on whether
    the trip was taken on a weekday or a weekend.
    The function then calculates the total number of trips for each provider type by adding up the weekday and weekend counts, and creates a new Excel
    workbook using the openpyxl library. It writes the summary statistics to a new sheet in the workbook, including counts for each provider type on
    weekdays, weekends, and overall, as well as the percentage of each provider type for weekdays in a new sheet.
"""

import csv
from datetime import datetime
import openpyxl
import pandas as pd

# Function to check if a given date is a weekday (Monday to Friday)
def is_weekday(date):
    return date.weekday() < 5

# Function to check if a given date is a weekend (Saturday or Sunday)
def is_weekend(date):
    return date.weekday() >= 5

def total_trips():
    # Dictionary to keep track of the counts of different providers by day of the week
    weekday_counts = {"Primary": 0, "Broker": 0, "E-Hail": 0}
    weekend_counts = {"Primary": 0, "Broker": 0, "E-Hail": 0}

    with open('trips.csv') as file:
        reader = csv.DictReader(file)
        for row in reader:
            # Parse the date of the trip from the Tripdate column
            trip_date = datetime.strptime(row['Tripdate'], '%Y-%m-%d')
        
            if row['ProviderType'] == "Test Record":
                continue
            if row['ProviderType'] == "Deleted":
                continue
            if row['ProviderType'] == "Driver break record":
                continue
            if row['Outcome'] == "Fixed Route-Exclude":
                continue
            # Increment the count for the provider type of the trip, based on the day of the week
            if is_weekday(trip_date):
                weekday_counts[row['ProviderType']] += 1
            elif is_weekend(trip_date):
                weekend_counts[row['ProviderType']] += 1

    # Calculate the total counts by adding up the weekday and weekend counts
    total_counts = {provider: weekday_counts[provider] + weekend_counts[provider] for provider in weekday_counts}

    # Create a new Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "question_1_part_1"

    # Write the header row
    header = [""] + ["Primary", "Broker", "E-Hail", "Total"]
    sheet.append(header)

    # Write the weekday row
    weekday_row = ["Weekday (M to F)"] + [weekday_counts["Primary"], weekday_counts["Broker"], weekday_counts["E-Hail"], sum(weekday_counts.values())]
    sheet.append(weekday_row)

    # Write the weekend row
    weekend_row = ["Weekend (Sa & Su)"] + [weekend_counts["Primary"], weekend_counts["Broker"], weekend_counts["E-Hail"], sum(weekend_counts.values())]
    sheet.append(weekend_row)

    # Write the total row
    total_row = ["Total"] + [total_counts["Primary"], total_counts["Broker"], total_counts["E-Hail"], sum(total_counts.values())]
    sheet.append(total_row)

    # Save the workbook to a file
    workbook.save("SUMMARY_OUTPUTS.xlsx")

    # Load the workbook using openpyxl
    book = openpyxl.load_workbook('SUMMARY_OUTPUTS.xlsx')

    # Create a Pandas dataframe with the percentage of each provider type
    data = {
        "ProviderType": ["Primary", "Broker", "E-Hail"],
        "question_1_part_1": [
            weekday_counts["Primary"] / sum(weekday_counts.values()),
            weekday_counts["Broker"] / sum(weekday_counts.values()),
            weekday_counts["E-Hail"] / sum(weekday_counts.values())
        ]
    }

    df = pd.DataFrame(data)

    # Write the modified dataframe to a new sheet in the Excel file
    with pd.ExcelWriter('SUMMARY_OUTPUTS.xlsx', engine='openpyxl', mode='a') as writer:
        try:
            if 'question_1_part_1' in writer.book.sheetnames:
                # If the sheet already exists, remove it
                writer.book.remove(writer.book.sheetnames.index('question_1_part_1'))
                # Add the sheet with the modified dataframe
            df.to_excel(writer, sheet_name='question_1_part_1', index=False, header=True)
        except:
            pass
        worksheet = writer.sheets['question_1_part_1']
        for col in worksheet.columns:
            column = col[0].column_letter
            col_width = worksheet.column_dimensions[column].width
            worksheet.column_dimensions[column].width = 4 + col_width


